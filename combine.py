import os
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def combine_worksheets(input_folders, output_file, selected_worksheets):
    n = 1
    # Create a new workbook
    combined_workbook = Workbook()
    
    # Iterate through input folders
    for folder in input_folders:
        # Iterate through files in the input folder
        for filename in os.listdir(folder):
            if filename.endswith(".xlsx"):
                workbook = load_workbook(os.path.join(folder, filename), data_only=True)
                
                # Get the workbook name
                workbook_name = os.path.splitext(filename)[0]
                
                # Iterate through worksheets in the loaded workbook
                for sheet_name in workbook.sheetnames:
                    if sheet_name in selected_worksheets:
                        print(n)
                        # Get the worksheet
                        worksheet = workbook[sheet_name]
                        
                        # Create a new worksheet in the combined workbook with the same name as the original workbook
                        combined_workbook_sheet = combined_workbook.create_sheet(title=workbook_name)
                        
                        # Write worksheet name in the first cell of the new sheet
                        combined_workbook_sheet.cell(row=1, column=1, value=workbook_name)
                        
                        # Copy data from the original worksheet to the new worksheet
                        for row_idx, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
                            for col_idx, value in enumerate(row, start=2):
                                combined_workbook_sheet.cell(row=row_idx, column=col_idx, value=value)
                        n = n + 1
    
    # Save the combined workbook
    combined_workbook.save(output_file)

# # # Example usage
input_folders = ["2023", "2022","2021","2019","2018","2017","2016","2015","2014","2013","2012","2010","2009","2008","2007","2005-2006","2004-2005","2003-2004","2002-2003","2001-2002","2000-2001"]  # Add paths to the input folders
output_file = "50mBackWomenFinal.xlsx"
selected_worksheets = ["Women 50m Backstroke-Final"]  # Add the names of the worksheets you want to select

combine_worksheets(input_folders, output_file, selected_worksheets)



# output_file = "200mFlyMenFinal.xlsx"
# selected_worksheets = ["Men 200m Butterfly-Final"]  # Add the names of the worksheets you want to select



# output_file = "400mFreeMenHeats.xlsx"
# selected_worksheets = ["Men 400m Freestyle-Heats Summar"]  # Add the names of the worksheets you want to select

# combine_worksheets(input_folders, output_file, selected_worksheets)


# output_file = "800mFreeMenFinal.xlsx"
# selected_worksheets = ["Men 800m Freestyle-Finals Summa"]  # Add the names of the worksheets you want to select

# combine_worksheets(input_folders, output_file, selected_worksheets)

# output_file = "1500mFreeMenFinal.xlsx"
# selected_worksheets = ["Men 1500m Freestyle-Finals Summ"]  # Add the names of the worksheets you want to select

# combine_worksheets(input_folders, output_file, selected_worksheets)
