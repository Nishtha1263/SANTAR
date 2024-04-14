import openpyxl

def merge_workbooks(first_workbook, second_workbook):
    # Create a new workbook
    new_wb = openpyxl.Workbook()
    
    # Load the first workbook
    first_wb = openpyxl.load_workbook(first_workbook)
    
    # Copy data from the first workbook into the new workbook
    for sheet_name in first_wb.sheetnames:
        first_sheet = first_wb[sheet_name]
        new_sheet = new_wb.create_sheet(title=sheet_name)
        for row in first_sheet.iter_rows(values_only=True):
            new_sheet.append(row)
    
    # Load the second workbook
    second_wb = openpyxl.load_workbook(second_workbook)
    
    # Iterate through each worksheet in the second workbook
    for sheet_name in second_wb.sheetnames:
        # Check if the worksheet exists in the new workbook. If not, create it.
        if sheet_name not in new_wb.sheetnames:
            new_sheet = new_wb.create_sheet(title=sheet_name)
        else:
            new_sheet = new_wb[sheet_name]
        
        second_sheet = second_wb[sheet_name]
        
        # Find the maximum column index in the new worksheet
        max_column_index = new_sheet.max_column
        
        # Append data from the second workbook to the right of the existing data
        for row_index, row in enumerate(second_sheet.iter_rows(values_only=True), start=1):
            for col_index, value in enumerate(row, start=1):
                new_sheet.cell(row=row_index, column=max_column_index + col_index, value=value)
    
    # Save the new workbook
    merged_workbook_name = "HvF.xlsx"
    new_wb.save(merged_workbook_name)
    print(f"Merged workbook saved as '{merged_workbook_name}'.")

# Example usage:
first_workbook = "50mBackWomenFinal.xlsx"
second_workbook = "50mBackWomenHeats.xlsx"

merge_workbooks(first_workbook, second_workbook)
